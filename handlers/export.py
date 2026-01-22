import pandas as pd
import io
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes
from handlers.base import BaseHandler
from config import UserRole, TaskStatus
from utils import format_datetime, get_status_emoji, get_priority_emoji
import logging

logger = logging.getLogger(__name__)

class ExportHandler(BaseHandler):
    def __init__(self, db):
        super().__init__(db)
        self.user_states = {}
    
    async def handle_export_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Export menyusi"""
        user = self.get_user(update)
        
        if not self.check_permission(user, [UserRole.SUPER_ADMIN, UserRole.ADMIN]):
            await self.send_message(update, context, "âŒ Bu funksiya faqat adminlar uchun!")
            return
        
        text = """
ğŸ“¤ <b>Ma'lumotlarni eksport qilish</b>

Qanday formatda eksport qilmoqchisiz?
        """
        
        keyboard = [
            [InlineKeyboardButton("ğŸ“Š XLSX (Excel)", callback_data="export_xlsx")],
            [InlineKeyboardButton("ğŸ”™ Orqaga", callback_data="main_menu")]
        ]
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await self.send_message(update, context, text, reply_markup)
    
    async def handle_export_xlsx(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """XLSX export"""
        user = self.get_user(update)
        
        if not self.check_permission(user, [UserRole.SUPER_ADMIN, UserRole.ADMIN]):
            await self.send_message(update, context, "âŒ Bu funksiya faqat adminlar uchun!")
            return
        
        # Filtr menyusini ko'rsatish
        text = """
ğŸ“Š <b>XLSX eksport</b>

Qanday ma'lumotlarni eksport qilmoqchisiz?
        """
        
        keyboard = [
            [InlineKeyboardButton("ğŸ“‹ Barcha vazifalar", callback_data="export_all_xlsx")],
            [InlineKeyboardButton("ğŸ‘¤ Ishchi bo'yicha", callback_data="export_user_xlsx")],
            [InlineKeyboardButton("ğŸ”™ Orqaga", callback_data="export_menu")]
        ]
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await self.send_message(update, context, text, reply_markup)
    
    
    async def handle_export_all(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Barcha vazifalarni eksport qilish"""
        user = self.get_user(update)
        data = update.callback_query.data
        export_type = data.split('_')[2]  # xlsx (export_all_xlsx -> [export, all, xlsx])
        
        logger.info(f"Export all called: user={user['id']}, data={data}, export_type={export_type}")
        
        if not self.check_permission(user, [UserRole.SUPER_ADMIN, UserRole.ADMIN]):
            await self.send_message(update, context, "âŒ Bu funksiya faqat adminlar uchun!")
            return
        
        try:
            # Barcha vazifalarni olish
            logger.info("Getting all tasks...")
            tasks = self.get_all_tasks_with_details()
            logger.info(f"Found {len(tasks)} tasks")
            
            if not tasks:
                await self.send_message(update, context, "ğŸ“ Eksport qilish uchun ma'lumotlar yo'q!")
                return
            
            # Export qilish
            logger.info("Creating XLSX export...")
            file_data = self.create_xlsx_export(tasks)
            filename = f"vazifalar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            logger.info(f"Created file: {filename}, size: {len(file_data)} bytes")
            
            # Faylni yuborish
            logger.info("Sending file...")
            await self.send_file(update, context, file_data, filename, export_type, task_count=len(tasks))
            logger.info("File sent successfully!")
            
            # Audit log
            self.db.add_audit_log(user['id'], 'EXPORT_ALL', f"Barcha vazifalar {export_type.upper()} formatida eksport qilindi")
            
        except Exception as e:
            logger.error(f"Export qilishda xatolik: {e}", exc_info=True)
            await self.send_message(update, context, "âŒ Export qilishda xatolik yuz berdi!")
    
    async def handle_export_date(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Sana bo'yicha eksport"""
        user = self.get_user(update)
        data = update.callback_query.data
        export_type = data.split('_')[2]  # xlsx (export_date_xlsx -> [export, date, xlsx])
        
        if not self.check_permission(user, [UserRole.SUPER_ADMIN, UserRole.ADMIN]):
            await self.send_message(update, context, "âŒ Bu funksiya faqat adminlar uchun!")
            return
        
        self.user_states[user['id']] = f'export_date_{export_type}'
        
        text = f"""
ğŸ“… <b>Sana bo'yicha eksport ({export_type.upper()})</b>

Boshlanish sanasini yuboring (DD.MM.YYYY formatida):
Masalan: 01.01.2024
        """
        
        reply_markup = self.create_back_button("export_menu")
        await self.send_message(update, context, text, reply_markup)
    
    async def handle_export_user(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Ishchi bo'yicha eksport"""
        user = self.get_user(update)
        data = update.callback_query.data
        export_type = data.split('_')[2]  # xlsx (export_date_xlsx -> [export, date, xlsx])
        
        if not self.check_permission(user, [UserRole.SUPER_ADMIN, UserRole.ADMIN]):
            await self.send_message(update, context, "âŒ Bu funksiya faqat adminlar uchun!")
            return
        
        # Ishchilar ro'yxatini olish
        workers = self.db.get_active_users()
        workers = [w for w in workers if w['role'] == UserRole.WORKER]
        
        if not workers:
            await self.send_message(update, context, "âŒ Ishchilar topilmadi!")
            return
        
        text = f"""
ğŸ‘¤ <b>Ishchi bo'yicha eksport ({export_type.upper()})</b>

Ishchini tanlang:
        """
        
        keyboard = []
        for worker in workers:
            keyboard.append([
                InlineKeyboardButton(
                    f"ğŸ‘· {worker['full_name']}", 
                    callback_data=f"export_worker_{worker['id']}_{export_type}"
                )
            ])
        
        keyboard.append([InlineKeyboardButton("ğŸ”™ Orqaga", callback_data="export_menu")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await self.send_message(update, context, text, reply_markup)
    
    async def handle_export_status(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Status bo'yicha eksport"""
        user = self.get_user(update)
        data = update.callback_query.data
        export_type = data.split('_')[2]  # xlsx (export_date_xlsx -> [export, date, xlsx])
        
        if not self.check_permission(user, [UserRole.SUPER_ADMIN, UserRole.ADMIN]):
            await self.send_message(update, context, "âŒ Bu funksiya faqat adminlar uchun!")
            return
        
        text = f"""
ğŸ“Š <b>Status bo'yicha eksport ({export_type.upper()})</b>

Statusni tanlang:
        """
        
        keyboard = [
            [InlineKeyboardButton("ğŸ“… SCHEDULED", callback_data=f"export_status_SCHEDULED_{export_type}")],
            [InlineKeyboardButton("ğŸ”„ IN_PROGRESS", callback_data=f"export_status_IN_PROGRESS_{export_type}")],
            [InlineKeyboardButton("â³ WAITING_APPROVAL", callback_data=f"export_status_WAITING_APPROVAL_{export_type}")],
            [InlineKeyboardButton("âœ… DONE", callback_data=f"export_status_DONE_{export_type}")],
            [InlineKeyboardButton("âŒ REJECTED", callback_data=f"export_status_REJECTED_{export_type}")],
            [InlineKeyboardButton("ğŸš¨ OVERDUE", callback_data=f"export_status_OVERDUE_{export_type}")],
            [InlineKeyboardButton("ğŸ”™ Orqaga", callback_data="export_menu")]
        ]
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await self.send_message(update, context, text, reply_markup)
    
    def get_all_tasks_with_details(self) -> List[Dict[str, Any]]:
        """Barcha vazifalarni tafsilotlari bilan olish"""
        query = """
            SELECT 
                t.id,
                t.title,
                t.description,
                t.status,
                t.priority,
                t.start_at,
                t.deadline,
                t.completed_at,
                t.approved_at,
                t.rejected_at,
                t.is_penalized,
                t.penalty_amount,
                t.created_at,
                u1.full_name as creator_name,
                u1.phone as creator_phone,
                u2.full_name as assigned_name,
                u2.phone as assigned_phone,
                u3.full_name as approver_name,
                u4.full_name as rejector_name
            FROM tasks t
            LEFT JOIN users u1 ON t.created_by = u1.id
            LEFT JOIN users u2 ON t.assigned_to = u2.id
            LEFT JOIN users u3 ON t.approved_by = u3.id
            LEFT JOIN users u4 ON t.rejected_by = u4.id
            ORDER BY t.created_at DESC
        """
        tasks = self.db.execute_query(query)
        
        # Har bir vazifa uchun deadline uzaytirish ma'lumotlarini qo'shish
        for task in tasks:
            try:
                extensions = self.db.get_task_deadline_extensions(task['id'])
                if extensions and len(extensions) > 0:
                    # Oxirgi uzaytirish ma'lumotini olish
                    latest_extension = extensions[0]
                    task['original_deadline'] = latest_extension.get('old_deadline', task['deadline'])
                    task['deadline_extended_by'] = latest_extension.get('extended_by_name', 'Nomalum')
                    task['deadline_extension_reason'] = latest_extension.get('reason', 'Yoq')
                    task['deadline_extension_hours'] = latest_extension.get('extension_hours', 0)
                else:
                    task['original_deadline'] = task['deadline']
                    task['deadline_extended_by'] = 'Uzaytirilmagan'
                    task['deadline_extension_reason'] = 'Yoq'
                    task['deadline_extension_hours'] = 0
            except Exception as e:
                # Xatolik bo'lsa, default qiymatlarni qo'yish
                task['original_deadline'] = task['deadline']
                task['deadline_extended_by'] = 'Uzaytirilmagan'
                task['deadline_extension_reason'] = 'Yoq'
                task['deadline_extension_hours'] = 0
        
        return tasks
    
    def get_tasks_by_date_range(self, start_date: str, end_date: str) -> List[Dict[str, Any]]:
        """Sana oralig'ida vazifalarni olish"""
        query = """
            SELECT 
                t.id,
                t.title,
                t.description,
                t.status,
                t.priority,
                t.start_at,
                t.deadline,
                t.completed_at,
                t.approved_at,
                t.is_penalized,
                t.penalty_amount,
                t.created_at,
                u1.full_name as creator_name,
                u2.full_name as assigned_name,
                u3.full_name as approver_name
            FROM tasks t
            LEFT JOIN users u1 ON t.created_by = u1.id
            LEFT JOIN users u2 ON t.assigned_to = u2.id
            LEFT JOIN users u3 ON t.approved_by = u3.id
            WHERE DATE(t.created_at) BETWEEN ? AND ?
            ORDER BY t.created_at DESC
        """
        return self.db.execute_query(query, (start_date, end_date))
    
    def get_tasks_by_user(self, user_id: int) -> List[Dict[str, Any]]:
        """Foydalanuvchi bo'yicha vazifalarni olish"""
        query = """
            SELECT 
                t.id,
                t.title,
                t.description,
                t.status,
                t.priority,
                t.start_at,
                t.deadline,
                t.completed_at,
                t.approved_at,
                t.rejected_at,
                t.is_penalized,
                t.penalty_amount,
                t.created_at,
                u1.full_name as creator_name,
                u1.phone as creator_phone,
                u2.full_name as assigned_name,
                u2.phone as assigned_phone,
                u3.full_name as approver_name,
                u4.full_name as rejector_name
            FROM tasks t
            LEFT JOIN users u1 ON t.created_by = u1.id
            LEFT JOIN users u2 ON t.assigned_to = u2.id
            LEFT JOIN users u3 ON t.approved_by = u3.id
            LEFT JOIN users u4 ON t.rejected_by = u4.id
            WHERE t.assigned_to = ?
            ORDER BY t.created_at DESC
        """
        tasks = self.db.execute_query(query, (user_id,))
        
        # Har bir vazifa uchun deadline uzaytirish ma'lumotlarini qo'shish
        for task in tasks:
            try:
                extensions = self.db.get_task_deadline_extensions(task['id'])
                if extensions and len(extensions) > 0:
                    latest_extension = extensions[0]
                    task['original_deadline'] = latest_extension.get('old_deadline', task['deadline'])
                    task['deadline_extended_by'] = latest_extension.get('extended_by_name', 'Nomalum')
                    task['deadline_extension_reason'] = latest_extension.get('reason', 'Yoq')
                    task['deadline_extension_hours'] = latest_extension.get('extension_hours', 0)
                else:
                    task['original_deadline'] = task['deadline']
                    task['deadline_extended_by'] = 'Uzaytirilmagan'
                    task['deadline_extension_reason'] = 'Yoq'
                    task['deadline_extension_hours'] = 0
            except Exception as e:
                task['original_deadline'] = task['deadline']
                task['deadline_extended_by'] = 'Uzaytirilmagan'
                task['deadline_extension_reason'] = 'Yoq'
                task['deadline_extension_hours'] = 0
        
        return tasks
    
    def get_tasks_by_status(self, status: str) -> List[Dict[str, Any]]:
        """Status bo'yicha vazifalarni olish"""
        query = """
            SELECT 
                t.id,
                t.title,
                t.description,
                t.status,
                t.priority,
                t.start_at,
                t.deadline,
                t.completed_at,
                t.approved_at,
                t.is_penalized,
                t.penalty_amount,
                t.created_at,
                u1.full_name as creator_name,
                u2.full_name as assigned_name,
                u3.full_name as approver_name
            FROM tasks t
            LEFT JOIN users u1 ON t.created_by = u1.id
            LEFT JOIN users u2 ON t.assigned_to = u2.id
            LEFT JOIN users u3 ON t.approved_by = u3.id
            WHERE t.status = ?
            ORDER BY t.created_at DESC
        """
        return self.db.execute_query(query, (status,))
    
    def create_xlsx_export(self, tasks: List[Dict[str, Any]]) -> bytes:
        """XLSX fayl yaratish"""
        # Ma'lumotlarni tayyorlash
        data = []
        for task in tasks:
            # Status emoji qo'shish
            status_emoji = get_status_emoji(task['status'])
            
            # Priority emoji qo'shish
            priority_emoji = get_priority_emoji(task['priority'])
            
            data.append({
                'ID': task['id'],
                'Sarlavha': task['title'],
                'Tavsif': task['description'] or 'Tavsif yoq',
                'Status': f"{status_emoji} {task['status']}",
                'Ustuvorlik': f"{priority_emoji} {task['priority']}",
                'Boshlanish vaqti': format_datetime(task['start_at']),
                'Original deadline': format_datetime(task.get('original_deadline', task['deadline'])),
                'Joriy deadline': format_datetime(task['deadline']),
                'Deadline uzaytirgan': task.get('deadline_extended_by', 'Uzaytirilmagan'),
                'Uzaytirish sababi': task.get('deadline_extension_reason', 'Yoq'),
                'Uzaytirish soati': f"{task.get('deadline_extension_hours', 0)} soat" if task.get('deadline_extension_hours', 0) > 0 else 'Yoq',
                'Ishchi tugatgan vaqt': format_datetime(task['completed_at']) if task['completed_at'] else 'Tugatilmagan',
                'Admin tasdiqlagan vaqt': format_datetime(task['approved_at']) if task['approved_at'] else 'Tasdiqlanmagan',
                'Rad etilgan vaqt': format_datetime(task.get('rejected_at')) if task.get('rejected_at') else 'Rad etilmagan',
                'Yaratuvchi': task['creator_name'] or 'Nomalum',
                'Yaratuvchi telefon': task['creator_phone'] or 'Kiritilmagan',
                'Ishchi': task['assigned_name'] or 'Tayinlanmagan',
                'Ishchi telefon': task['assigned_phone'] or 'Kiritilmagan',
                'Tasdiqlovchi': task['approver_name'] or 'Tasdiqlanmagan',
                'Rad etuvchi': task['rejector_name'] or 'Rad etilmagan',
                'Jarima': 'ğŸ’° Ha' if task['is_penalized'] else 'âœ… Yoq',
                'Jarima miqdori': f"{task['penalty_amount']:,} UZS" if task['penalty_amount'] > 0 else '0 UZS',
                'Yaratilgan': format_datetime(task['created_at'])
            })
        
        # DataFrame yaratish
        df = pd.DataFrame(data)
        
        # Excel fayl yaratish
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Vazifalar', index=False)
            
            # Ustun kengliklarini o'rnatish
            worksheet = writer.sheets['Vazifalar']
            
            # Ustun kengliklari
            column_widths = {
                'A': 35,  # ID
                'B': 20,  # Sarlavha
                'C': 30,  # Tavsif
                'D': 15,  # Status
                'E': 15,  # Ustuvorlik
                'F': 18,  # Boshlanish vaqti
                'G': 18,  # Original deadline
                'H': 18,  # Joriy deadline
                'I': 20,  # Deadline uzaytirgan
                'J': 25,  # Uzaytirish sababi
                'K': 15,  # Uzaytirish soati
                'L': 20,  # Ishchi tugatgan vaqt
                'M': 20,  # Admin tasdiqlagan vaqt
                'N': 20,  # Rad etilgan vaqt
                'O': 20,  # Yaratuvchi
                'P': 18,  # Yaratuvchi telefon
                'Q': 20,  # Ishchi
                'R': 18,  # Ishchi telefon
                'S': 20,  # Tasdiqlovchi
                'T': 20,  # Rad etuvchi
                'U': 12,  # Jarima
                'V': 15,  # Jarima miqdori
                'W': 18   # Yaratilgan
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Sarlavha qatorini formatlash
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Ma'lumotlar qatorlarini formatlash
            data_alignment = Alignment(horizontal="left", vertical="center")
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = data_alignment
        
        output.seek(0)
        return output.getvalue()
    
    
    async def send_file(self, update: Update, context: ContextTypes.DEFAULT_TYPE, 
                       file_data: bytes, filename: str, file_type: str, task_count: int = None):
        """Faylni yuborish"""
        try:
            logger.info(f"send_file called: filename={filename}, file_type={file_type}, data_size={len(file_data)}")
            if file_type == 'xlsx':
                from telegram import InputFile
                from utils import format_datetime
                
                # Fayl hajmini hisoblash
                file_size_kb = len(file_data) / 1024
                file_size_mb = file_size_kb / 1024
                
                if file_size_mb >= 1:
                    size_text = f"{file_size_mb:.2f} MB"
                else:
                    size_text = f"{file_size_kb:.2f} KB"
                
                # Hozirgi vaqt
                current_time = format_datetime(datetime.now())
                
                # Chiroyli caption yaratish
                caption = f"""
âœ… <b>Eksport muvaffaqiyatli yakunlandi!</b>

â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”
ğŸ“Š <b>Ma'lumotlar:</b>
â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”

ğŸ“ <b>Fayl nomi:</b>
<code>{filename}</code>

ğŸ“Š <b>Format:</b> XLSX (Excel)

ğŸ’¾ <b>Fayl hajmi:</b> {size_text}
"""
                
                # Agar vazifalar soni berilgan bo'lsa
                if task_count is not None:
                    caption += f"ğŸ“‹ <b>Vazifalar soni:</b> {task_count} ta\n"
                
                caption += f"""
ğŸ• <b>Eksport vaqti:</b> {current_time}

â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”
ğŸ’¡ <b>Eslatma:</b> Faylni ochish uchun yuklab oling
â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”
"""
                
                document = InputFile(io.BytesIO(file_data), filename=filename)
                logger.info(f"Sending document to chat_id={update.effective_chat.id}")
                await context.bot.send_document(
                    chat_id=update.effective_chat.id,
                    document=document,
                    caption=caption,
                    parse_mode='HTML'
                )
                logger.info("Document sent successfully!")
        except Exception as e:
            logger.error(f"Fayl yuborishda xatolik: {e}", exc_info=True)
            await self.send_message(update, context, "âŒ Fayl yuborishda xatolik yuz berdi!")
